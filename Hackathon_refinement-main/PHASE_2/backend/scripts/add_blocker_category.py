from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

WORKBOOK_PATH = "PHASE_2/INPUT/TIO2_Sprint_Intelligence_v5_final.xlsx"

CATEGORY_MAP = {
    "BLK-001": "External Team Dependency",
    "BLK-002": "Lab Issue",
    "BLK-003": "Hardware / Procurement",
    "BLK-004": "External Team Dependency",
    "BLK-005": "Lab Issue",
}

wb = load_workbook(WORKBOOK_PATH)
ws = wb["Blockers"]

header_row = None
for row in ws.iter_rows():
    for cell in row:
        if cell.value == "Blocker ID":
            header_row = cell.row
            break
    if header_row:
        break

if header_row is None:
    raise ValueError("Blocker ID header not found")

header_cell = ws.cell(row=header_row, column=11, value="Category")
header_cell.font = Font(bold=True)
header_cell.alignment = Alignment(horizontal="center")

blocker_id_col = None
for cell in ws[header_row]:
    if cell.value == "Blocker ID":
        blocker_id_col = cell.column
        break

if blocker_id_col is None:
    raise ValueError("Blocker ID column not found")

for row in ws.iter_rows(min_row=header_row + 1):
    blocker_id = row[blocker_id_col - 1].value
    if not blocker_id:
        continue
    category = CATEGORY_MAP.get(str(blocker_id).strip(), "Other")
    cat_cell = ws.cell(row=row[0].row, column=11, value=category)
    cat_cell.alignment = Alignment(horizontal="left")

wb.save(WORKBOOK_PATH)
print("Done. Verify output:")
for row in ws.iter_rows(min_row=header_row, max_row=header_row+6, values_only=True):
    print(row)
